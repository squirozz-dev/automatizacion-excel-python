import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Crear libro y hoja
wb = Workbook()
ws = wb.active
ws.title = "Reporte de Ventas"

# Estilos reutilizables
bold_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
encabezados = ["Producto", "Cantidad", "Precio Unitario", "Total"]
ws.append(encabezados)

# Aplicar estilo a encabezados
for cell in ws[1]:
    cell.font = bold_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Datos
ventas = [
    ["Camiseta", 3, 25000],
    ["Pantalón", 2, 40000],
    ["Zapatos", 1, 120000],
    ["Gorra", 5, 10000],
]

# Agregar datos y bordes
for fila in ventas:
    producto, cantidad, precio = fila
    total = cantidad * precio
    ws.append([producto, cantidad, precio, total])

# Aplicar bordes a las filas de datos
for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border

# Agregar total general
fila_total = len(ventas) + 2
ws[f"A{fila_total}"] = "Total general:"
ws[f"A{fila_total}"].font = Font(bold=True, color="006100")
ws[f"D{fila_total}"] = f"=SUM(D2:D{fila_total - 1})"
ws[f"D{fila_total}"].font = Font(bold=True, color="006100")

# Bordes y alineación en fila final
for col in ["A", "B", "C", "D"]:
    ws[f"{col}{fila_total}"].border = thin_border
    ws[f"{col}{fila_total}"].alignment = Alignment(horizontal="center")

# Ajuste automático de columnas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # A, B, C...
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column].width = max_length + 2

# Guardar
fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
ruta = os.path.join(os.path.expanduser("~"), "Desktop", f"Reporte_Ventas_Pro_{fecha}.xlsx")
wb.save(ruta)

print(f"Archivo guardado exitosamente en: {ruta}")
